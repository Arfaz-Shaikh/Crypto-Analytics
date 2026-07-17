import os
from datetime import datetime
import os
import sys
print("CWD:", os.getcwd())
print("sys.path:", sys.path)

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import os
import sys
print("Current Working Directory:", os.getcwd())
print("Python Executable:", sys.executable)
print("Python Path:", sys.path)

from config.database import engine

# -----------------------------
# SQL Queries
# -----------------------------

EXECUTIVE_QUERY = """
SELECT
    COUNT(*) AS total_coins,
    SUM(market_cap) AS total_market_cap,
    AVG(current_price) AS average_price
FROM crypto.vw_latest_market_snapshot;
"""

TOP_MARKET_CAP_QUERY = """
SELECT
    coin_name,
    market_cap,
    current_price
FROM crypto.vw_latest_market_snapshot
ORDER BY market_cap DESC
LIMIT 10;
"""

TOP_VOLUME_QUERY = """
SELECT
    coin_name,
    total_volume
FROM crypto.vw_latest_market_snapshot
ORDER BY total_volume DESC
LIMIT 10;
"""

TOP_GAINERS_QUERY = """
SELECT
    coin_name,
    price_change_percentage_24h
FROM crypto.vw_latest_market_snapshot
ORDER BY price_change_percentage_24h DESC
LIMIT 10;
"""

ETL_AUDIT_QUERY = """
SELECT *
FROM crypto.etl_audit
ORDER BY run_start DESC
LIMIT 20;
"""


# -----------------------------
# Read SQL Query
# -----------------------------

def run_query(query):
    return pd.read_sql(query, engine)


# -----------------------------
# Create Excel Report
# -----------------------------

def create_report():

    output_folder = "reports/excel"
    os.makedirs(output_folder, exist_ok=True)

    filename = f"Crypto_Report_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    filepath = os.path.join(output_folder, filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:

        run_query(EXECUTIVE_QUERY).to_excel(
            writer,
            sheet_name="Executive Summary",
            index=False
        )

        run_query(TOP_MARKET_CAP_QUERY).to_excel(
            writer,
            sheet_name="Top Market Cap",
            index=False
        )

        run_query(TOP_VOLUME_QUERY).to_excel(
            writer,
            sheet_name="Top Volume",
            index=False
        )

        run_query(TOP_GAINERS_QUERY).to_excel(
            writer,
            sheet_name="Top Gainers",
            index=False
        )

        run_query(ETL_AUDIT_QUERY).to_excel(
            writer,
            sheet_name="ETL Audit",
            index=False
        )

    return filepath


# -----------------------------
# Format Workbook
# -----------------------------

def format_workbook(filepath):

    wb = load_workbook(filepath)

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    for sheet in wb.sheetnames:

        ws = wb[sheet]

        for cell in ws[1]:

            cell.fill = header_fill
            cell.font = header_font

        for column in ws.columns:

            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:

                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            ws.column_dimensions[column_letter].width = max_length + 3

    wb.save(filepath)


# -----------------------------
# Main
# -----------------------------

def main():

    print("Generating Excel Report...")

    filepath = create_report()

    format_workbook(filepath)

    print("Excel Report Created Successfully!")

    print(filepath)


if __name__ == "__main__":
    main()